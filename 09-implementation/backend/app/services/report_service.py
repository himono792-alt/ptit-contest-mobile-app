"""Reports / aggregate stats (GV-07, BCN-03, BCN-05, AD-05)."""

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, status
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.certificate import IssuedCertificate
from app.models.contest import Contest, ContestRound
from app.models.entry import ContestEntry, Team, TeamMember
from app.models.enums import (
    ContestStatus,
    EntryType,
    RegistrationStatus,
    RoleCode,
    SubmissionStatus,
    UserStatus,
)
from app.models.identity import AppUser, Role, UserRole
from app.models.judging import ContestResult, RoundResult
from app.models.master_data import (
    DepartmentHead,
    Faculty,
    Student,
    StudentDirectory,
)
from app.models.review import ContestReview
from app.models.submission import Submission


# ---------- Helpers ----------

async def _ensure_organizer_or_admin(user, contest):
    if "ADMIN" in user.role_codes:
        return
    if "ORGANIZER" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ORGANIZER hoặc ADMIN")
    if contest.created_by != user.user_id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Không phải BTC contest này")


async def _ensure_hod_or_admin(db: AsyncSession, user) -> int | None:
    """Trả faculty_id để filter (None = ADMIN xem tất)."""
    if "ADMIN" in user.role_codes:
        return None
    if "HOD" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần HOD hoặc ADMIN")
    dh = (await db.execute(
        select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
    )).scalar_one_or_none()
    if dh is None:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "User chưa có profile HOD")
    return dh.faculty_id


# ============================================================
# GV-07 contest stats
# ============================================================

async def contest_stats(db: AsyncSession, user, contest_id: int) -> dict:
    contest = await db.get(Contest, contest_id)
    if contest is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contest not found")
    await _ensure_organizer_or_admin(user, contest)

    # Entries by status
    entries_by_status = (await db.execute(
        select(
            ContestEntry.registration_status,
            func.count().label("c"),
        )
        .where(ContestEntry.contest_id == contest_id)
        .group_by(ContestEntry.registration_status)
    )).all()
    by_status = {row.registration_status: row.c for row in entries_by_status}
    total_entries = sum(by_status.values())

    # Submissions
    sub_stats = (await db.execute(
        select(
            func.count().label("total"),
            func.sum(case((Submission.status == SubmissionStatus.SUBMITTED, 1), else_=0)).label("submitted"),
            func.sum(case((Submission.status == SubmissionStatus.LATE, 1), else_=0)).label("late"),
        )
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .where(ContestRound.contest_id == contest_id)
    )).one()

    # Rounds
    rounds_count = (await db.execute(
        select(func.count()).select_from(ContestRound).where(ContestRound.contest_id == contest_id)
    )).scalar_one()

    rounds_with_results = (await db.execute(
        select(func.count(func.distinct(RoundResult.round_id)))
        .select_from(RoundResult)
        .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
        .where(ContestRound.contest_id == contest_id)
    )).scalar_one()

    # Final score stats (avg + pass rate top 50%)
    final_stats = (await db.execute(
        select(
            func.avg(ContestResult.final_score).label("avg_score"),
            func.count().label("c"),
        )
        .where(
            ContestResult.contest_id == contest_id,
            ContestResult.final_score.is_not(None),
        )
    )).one()
    avg_score = final_stats.avg_score
    total_results = final_stats.c

    pass_rate = None
    if total_results > 0:
        # Top 50% = rank <= total/2
        top_half = (await db.execute(
            select(func.count())
            .select_from(ContestResult)
            .where(
                ContestResult.contest_id == contest_id,
                ContestResult.rank_no.is_not(None),
                ContestResult.rank_no <= max(1, total_results // 2),
            )
        )).scalar_one()
        pass_rate = Decimal(top_half * 100) / Decimal(total_results)

    # Reviews
    rev_stats = (await db.execute(
        select(
            func.avg(ContestReview.rating).label("avg"),
            func.count().label("c"),
        )
        .where(
            ContestReview.contest_id == contest_id,
            ContestReview.is_visible.is_(True),
        )
    )).one()

    return {
        "contest_id": contest_id,
        "contest_title": contest.title,
        "contest_status": contest.status.value,
        "total_entries": total_entries,
        "approved_entries": by_status.get(RegistrationStatus.APPROVED, 0),
        "pending_entries": by_status.get(RegistrationStatus.PENDING, 0),
        "rejected_entries": by_status.get(RegistrationStatus.REJECTED, 0),
        "cancelled_entries": by_status.get(RegistrationStatus.CANCELLED, 0),
        "total_submissions": sub_stats.total or 0,
        "submitted_count": sub_stats.submitted or 0,
        "late_count": sub_stats.late or 0,
        "rounds_count": rounds_count,
        "rounds_with_results": rounds_with_results,
        "average_final_score": avg_score,
        "pass_rate_percent": pass_rate,
        "avg_review_rating": rev_stats.avg,
        "review_count": rev_stats.c,
    }


# ============================================================
# BCN-03 monitor
# ============================================================

async def monitor_contests(
    db: AsyncSession, user, faculty_filter: int | None = None
) -> list[dict]:
    user_faculty = await _ensure_hod_or_admin(db, user)
    if user_faculty is not None:
        faculty_filter = user_faculty  # HOD chỉ xem khoa mình

    base = select(Contest).order_by(Contest.start_at.desc())
    if faculty_filter is not None:
        base = base.where(Contest.host_faculty_id == faculty_filter)
    contests = (await db.execute(base)).scalars().all()

    items: list[dict] = []
    for c in contests:
        # Count entries
        entries_total = (await db.execute(
            select(func.count())
            .select_from(ContestEntry)
            .where(ContestEntry.contest_id == c.contest_id)
        )).scalar_one()
        entries_approved = (await db.execute(
            select(func.count())
            .select_from(ContestEntry)
            .where(
                ContestEntry.contest_id == c.contest_id,
                ContestEntry.registration_status == RegistrationStatus.APPROVED,
            )
        )).scalar_one()

        # Submissions
        sub_count = (await db.execute(
            select(func.count())
            .select_from(Submission)
            .join(ContestRound, ContestRound.round_id == Submission.round_id)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()

        # Rounds
        rounds_total = (await db.execute(
            select(func.count())
            .select_from(ContestRound)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()
        rounds_done = (await db.execute(
            select(func.count(func.distinct(RoundResult.round_id)))
            .select_from(RoundResult)
            .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()

        reg_pct = None
        if c.max_entries:
            reg_pct = round(entries_total / c.max_entries * 100, 1)

        sub_pct = None
        if entries_approved > 0:
            sub_pct = round(sub_count / entries_approved * 100, 1)

        judg_pct = None
        if rounds_total > 0:
            judg_pct = round(rounds_done / rounds_total * 100, 1)

        items.append({
            "contest_id": c.contest_id,
            "title": c.title,
            "slug": c.slug,
            "status": c.status.value,
            "start_at": c.start_at.isoformat() if c.start_at else None,
            "end_at": c.end_at.isoformat() if c.end_at else None,
            "registration_pct": reg_pct,
            "submission_pct": sub_pct,
            "judging_pct": judg_pct,
            "total_entries": entries_total,
            "total_submissions": sub_count,
        })
    return items


# ============================================================
# BCN-05 faculty summary
# ============================================================

async def faculty_summary(
    db: AsyncSession, user, faculty_id: int | None, year: int
) -> dict:
    user_faculty = await _ensure_hod_or_admin(db, user)
    target_faculty = user_faculty if user_faculty is not None else faculty_id
    if target_faculty is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Admin cần truyền faculty_id query param",
        )

    faculty = await db.get(Faculty, target_faculty)
    if faculty is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Faculty not found")

    # Year boundary
    from datetime import datetime, timezone
    year_start = datetime(year, 1, 1, tzinfo=timezone.utc)
    year_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    # Contests of faculty in year
    base_filter = [
        Contest.host_faculty_id == target_faculty,
        Contest.created_at >= year_start,
        Contest.created_at < year_end,
    ]

    total_contests = (await db.execute(
        select(func.count()).select_from(Contest).where(*base_filter)
    )).scalar_one()
    finished = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status == ContestStatus.FINISHED)
    )).scalar_one()
    ongoing = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status.in_([ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING]))
    )).scalar_one()
    draft_pending = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status.in_([ContestStatus.DRAFT, ContestStatus.PROPOSED, ContestStatus.REVISION_REQUESTED]))
    )).scalar_one()

    # Entries + unique students in faculty contests
    entries_total = (await db.execute(
        select(func.count())
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*base_filter)
    )).scalar_one()
    unique_students = (await db.execute(
        select(func.count(func.distinct(ContestEntry.student_id)))
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*base_filter, ContestEntry.student_id.is_not(None))
    )).scalar_one()

    # Awards
    awards_count = (await db.execute(
        select(func.count())
        .select_from(ContestResult)
        .join(Contest, Contest.contest_id == ContestResult.contest_id)
        .where(*base_filter, ContestResult.award_title.is_not(None))
    )).scalar_one()

    # Avg rating
    avg_rating = (await db.execute(
        select(func.avg(ContestReview.rating))
        .select_from(ContestReview)
        .join(Contest, Contest.contest_id == ContestReview.contest_id)
        .where(*base_filter, ContestReview.is_visible.is_(True))
    )).scalar_one()

    return {
        "faculty_id": target_faculty,
        "faculty_code": faculty.faculty_code,
        "faculty_name": faculty.faculty_name,
        "year": year,
        "total_contests": total_contests,
        "contests_finished": finished,
        "contests_ongoing": ongoing,
        "contests_draft_or_pending": draft_pending,
        "total_unique_students": unique_students,
        "total_entries": entries_total,
        "total_awards": awards_count,
        "avg_completion_rate": (
            Decimal(finished * 100) / Decimal(total_contests)
            if total_contests > 0 else None
        ),
        "avg_rating": avg_rating,
    }


# ============================================================
# AD-05 system summary
# ============================================================

async def system_summary(db: AsyncSession, user, year: int) -> dict:
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ADMIN")

    from datetime import datetime, timezone
    year_start = datetime(year, 1, 1, tzinfo=timezone.utc)
    year_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    # User counts
    total_users = (await db.execute(
        select(func.count()).select_from(AppUser)
        .where(AppUser.status != UserStatus.DELETED)
    )).scalar_one()

    counts_by_role = {}
    for rc in [RoleCode.STUDENT, RoleCode.ORGANIZER, RoleCode.JUDGE, RoleCode.HOD, RoleCode.ADMIN]:
        cnt = (await db.execute(
            select(func.count(func.distinct(AppUser.user_id)))
            .select_from(AppUser)
            .join(UserRole, UserRole.user_id == AppUser.user_id)
            .join(Role, Role.role_id == UserRole.role_id)
            .where(
                AppUser.status == UserStatus.ACTIVE,
                Role.role_code == rc,
            )
        )).scalar_one()
        counts_by_role[rc] = cnt

    # Contests
    contests_in_year = [Contest.created_at >= year_start, Contest.created_at < year_end]
    total_contests = (await db.execute(
        select(func.count()).select_from(Contest).where(*contests_in_year)
    )).scalar_one()
    published_or_after = (await db.execute(
        select(func.count()).select_from(Contest).where(
            *contests_in_year,
            Contest.status.in_([
                ContestStatus.PUBLISHED, ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED,
                ContestStatus.ONGOING, ContestStatus.FINISHED,
            ]),
        )
    )).scalar_one()
    finished = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*contests_in_year, Contest.status == ContestStatus.FINISHED)
    )).scalar_one()

    # Entries / submissions
    entries_total = (await db.execute(
        select(func.count())
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*contests_in_year)
    )).scalar_one()
    submissions_total = (await db.execute(
        select(func.count())
        .select_from(Submission)
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .join(Contest, Contest.contest_id == ContestRound.contest_id)
        .where(*contests_in_year)
    )).scalar_one()

    # Certificates
    certs_total = (await db.execute(
        select(func.count())
        .select_from(IssuedCertificate)
        .join(ContestResult, ContestResult.contest_result_id == IssuedCertificate.contest_result_id)
        .join(Contest, Contest.contest_id == ContestResult.contest_id)
        .where(*contests_in_year, IssuedCertificate.revoked_at.is_(None))
    )).scalar_one()

    # Reviews
    rev_stats = (await db.execute(
        select(
            func.count().label("c"),
            func.avg(ContestReview.rating).label("avg"),
        )
        .select_from(ContestReview)
        .join(Contest, Contest.contest_id == ContestReview.contest_id)
        .where(*contests_in_year, ContestReview.is_visible.is_(True))
    )).one()

    return {
        "year": year,
        "total_users": total_users,
        "students_active": counts_by_role[RoleCode.STUDENT],
        "organizers": counts_by_role[RoleCode.ORGANIZER],
        "judges": counts_by_role[RoleCode.JUDGE],
        "department_heads": counts_by_role[RoleCode.HOD],
        "admins": counts_by_role[RoleCode.ADMIN],
        "total_contests": total_contests,
        "contests_published_or_after": published_or_after,
        "contests_finished": finished,
        "total_entries": entries_total,
        "total_submissions": submissions_total,
        "total_certificates_issued": certs_total,
        "total_reviews": rev_stats.c or 0,
        "avg_review_rating": rev_stats.avg,
    }


# ============================================================
# Phase 2 sprint 1 step 3 (2026-05-06): Excel export contest results
# ============================================================

async def export_contest_results_xlsx(
    db: AsyncSession, user: AppUser, contest_id: int
) -> tuple[bytes, str]:
    """Render contest results to xlsx (4 sheets).

    Trả tuple (bytes_data, suggested_filename).
    Caller wrap vào StreamingResponse với content-type
    application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.

    Sheets:
    - Tổng quan: rank, SV/team, mã SV, total_score, award
    - Vòng: rank theo round, total_score, average_score, is_passed
    - Submissions: round_no, entry, status, submitted_at
    - Metadata: contest info, generated by, generated at

    Permission: BTC của contest, BCN, ADMIN.
    """
    # openpyxl import lazy để tránh load nếu module không dùng
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    contest = await db.get(Contest, contest_id)
    if contest is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contest not found")
    await _ensure_organizer_or_admin(user, contest)

    # ----- Fetch data -----
    # Sheet 1: Tổng quan — JOIN ContestResult → ContestEntry → Student/Team → AppUser
    overall_stmt = (
        select(ContestResult, ContestEntry)
        .join(ContestEntry, ContestEntry.entry_id == ContestResult.entry_id)
        .where(ContestResult.contest_id == contest_id)
        .order_by(ContestResult.rank_no.nulls_last())
    )
    overall_rows = (await db.execute(overall_stmt)).all()

    # Resolve display name + code cho từng entry
    async def _entry_display(entry: ContestEntry) -> tuple[str, str]:
        """Trả (display_name, identifier_code). Cho INDIVIDUAL: SV name + MSSV.
        Cho TEAM: team_name + team_id."""
        if entry.entry_type == EntryType.INDIVIDUAL and entry.student_id:
            stmt = (
                select(Student.student_code, AppUser.full_name)
                .join(AppUser, AppUser.user_id == Student.user_id)
                .where(Student.student_id == entry.student_id)
            )
            row = (await db.execute(stmt)).first()
            if row:
                return row.full_name, row.student_code
            return "(SV không xác định)", str(entry.student_id)
        if entry.entry_type == EntryType.TEAM and entry.team_id:
            team = await db.get(Team, entry.team_id)
            if team:
                return team.team_name, f"Team #{team.team_id}"
            return "(Team không xác định)", f"Team #{entry.team_id}"
        return "(Không xác định)", str(entry.entry_id)

    # Sheet 2: ContestRound + RoundResult
    rounds_stmt = (
        select(ContestRound)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no)
    )
    rounds = list((await db.execute(rounds_stmt)).scalars().all())

    round_results_stmt = (
        select(RoundResult, ContestRound)
        .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no, RoundResult.rank_no.nulls_last())
    )
    round_result_rows = (await db.execute(round_results_stmt)).all()

    # Sheet 3: Submissions
    submissions_stmt = (
        select(Submission, ContestRound)
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no, Submission.entry_id)
    )
    submission_rows = (await db.execute(submissions_stmt)).all()

    # ----- Build workbook -----
    wb = Workbook()
    # openpyxl tạo sẵn 1 sheet "Sheet" — em rename thành "Tổng quan"
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws2 = wb.create_sheet("Vòng - kết quả")
    ws3 = wb.create_sheet("Submissions")
    ws4 = wb.create_sheet("Metadata")

    # PTIT brand styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    def _write_headers(ws, headers: list[str]) -> None:
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        # Freeze pane header row
        ws.freeze_panes = "A2"
        # Autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _autosize(ws, max_width: int = 50) -> None:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_len:
                        max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)

    # ----- Sheet 1: Tổng quan -----
    _write_headers(ws1, ["Hạng", "Tên / Team", "Mã SV / Team ID", "Tổng điểm", "Giải thưởng", "BCN duyệt"])
    for ridx, (cr, entry) in enumerate(overall_rows, start=2):
        name, code = await _entry_display(entry)
        ws1.cell(row=ridx, column=1, value=cr.rank_no or "—")
        ws1.cell(row=ridx, column=2, value=name)
        ws1.cell(row=ridx, column=3, value=code)
        score_cell = ws1.cell(row=ridx, column=4,
                               value=float(cr.final_score) if cr.final_score is not None else None)
        score_cell.number_format = "0.00"
        ws1.cell(row=ridx, column=5, value=cr.award_title or "")
        ws1.cell(row=ridx, column=6, value=cr.bcn_approval_status.value if cr.bcn_approval_status else "")
    _autosize(ws1)

    # ----- Sheet 2: Vòng - kết quả -----
    _write_headers(ws2, ["Vòng", "Tên vòng", "Hạng", "Entry ID", "Tên / Team", "Tổng điểm", "Điểm TB", "Đậu?"])
    # Cache entry display để không query lặp
    entry_display_cache: dict[int, tuple[str, str]] = {}
    for cr, _entry in overall_rows:
        entry_display_cache[_entry.entry_id] = await _entry_display(_entry)

    row_idx = 2
    for rr, rnd in round_result_rows:
        # Lookup entry display (có thể không có trong overall_rows nếu entry bị eliminate giữa chừng)
        if rr.entry_id not in entry_display_cache:
            ent = await db.get(ContestEntry, rr.entry_id)
            if ent is not None:
                entry_display_cache[rr.entry_id] = await _entry_display(ent)
            else:
                entry_display_cache[rr.entry_id] = ("(deleted)", str(rr.entry_id))
        name, _code = entry_display_cache[rr.entry_id]
        ws2.cell(row=row_idx, column=1, value=rnd.round_no)
        ws2.cell(row=row_idx, column=2, value=rnd.round_name)
        ws2.cell(row=row_idx, column=3, value=rr.rank_no or "—")
        ws2.cell(row=row_idx, column=4, value=rr.entry_id)
        ws2.cell(row=row_idx, column=5, value=name)
        c1 = ws2.cell(row=row_idx, column=6, value=float(rr.total_score) if rr.total_score else None)
        c1.number_format = "0.00"
        c2 = ws2.cell(row=row_idx, column=7, value=float(rr.average_score) if rr.average_score else None)
        c2.number_format = "0.00"
        ws2.cell(row=row_idx, column=8, value="✓" if rr.is_passed else ("✗" if rr.is_passed is False else ""))
        row_idx += 1
    _autosize(ws2)

    # ----- Sheet 3: Submissions -----
    _write_headers(ws3, ["Vòng", "Tên vòng", "Submission ID", "Entry ID", "Tên / Team",
                         "Trạng thái", "Phiên bản", "Đã nộp lúc"])
    row_idx = 2
    for sub, rnd in submission_rows:
        if sub.entry_id not in entry_display_cache:
            ent = await db.get(ContestEntry, sub.entry_id)
            if ent is not None:
                entry_display_cache[sub.entry_id] = await _entry_display(ent)
            else:
                entry_display_cache[sub.entry_id] = ("(deleted)", str(sub.entry_id))
        name, _code = entry_display_cache[sub.entry_id]
        ws3.cell(row=row_idx, column=1, value=rnd.round_no)
        ws3.cell(row=row_idx, column=2, value=rnd.round_name)
        ws3.cell(row=row_idx, column=3, value=sub.submission_id)
        ws3.cell(row=row_idx, column=4, value=sub.entry_id)
        ws3.cell(row=row_idx, column=5, value=name)
        ws3.cell(row=row_idx, column=6, value=sub.status.value if sub.status else "")
        ws3.cell(row=row_idx, column=7, value=sub.current_version_no)
        ws3.cell(row=row_idx, column=8,
                 value=sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "")
        row_idx += 1
    _autosize(ws3)

    # ----- Sheet 4: Metadata -----
    _write_headers(ws4, ["Trường thông tin", "Giá trị"])
    metadata_rows = [
        ("Mã contest", contest.contest_id),
        ("Slug", contest.slug),
        ("Tên cuộc thi", contest.title),
        ("Trạng thái", contest.status.value if contest.status else ""),
        ("Hình thức tham gia", contest.participation_mode.value if contest.participation_mode else ""),
        ("Hình thức tổ chức", contest.delivery_mode.value if contest.delivery_mode else ""),
        ("Bắt đầu", contest.start_at.strftime("%Y-%m-%d %H:%M:%S") if contest.start_at else ""),
        ("Kết thúc", contest.end_at.strftime("%Y-%m-%d %H:%M:%S") if contest.end_at else ""),
        ("Số entries có kết quả", len(overall_rows)),
        ("Số rounds", len(rounds)),
        ("Số submissions", len(submission_rows)),
        ("Người xuất", f"{user.full_name} ({user.email})"),
        ("Thời gian xuất", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for ridx, (k, v) in enumerate(metadata_rows, start=2):
        ws4.cell(row=ridx, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=ridx, column=2, value=str(v) if v is not None else "")
    _autosize(ws4)

    # ----- Save to BytesIO + return bytes + filename -----
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"ket-qua-{contest.slug}-{today_str}.xlsx"

    return buf.read(), filename
