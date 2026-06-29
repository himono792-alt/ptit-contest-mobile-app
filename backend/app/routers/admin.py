"""Admin router — gộp 4 sub-router (users, master, configs, audit).

Endpoints:
  Users:        /api/admin/users/*  (AD-02)
  Master data:  /api/admin/faculties, /majors, /classes  (AD-03)
  Configs:      /api/admin/configs/*  (AD-04)
  Audit:        /api/admin/audit-logs, /reviews/*  (AD-06)
"""

from datetime import datetime
from typing import Annotated

from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import CurrentUser
from app.models.enums import RoleCode, UserStatus
from app.schemas.admin import (
    AcademicClassIn,
    AcademicClassOut,
    AdminUserCreateIn,
    AdminUserListItem,
    AdminUserListOut,
    AdminUserUpdateIn,
    AssignRolesIn,
    BulkUserStatusIn,
    BulkUserStatusOut,
    AuditLogListOut,
    AuditLogOut,
    BulkImportIn,
    BulkImportOut,
    ConfigOut,
    ConfigUpdateIn,
    FacultyIn,
    FacultyOut,
    MajorIn,
    MajorOut,
)
from app.schemas.review import BulkModerateIn, BulkModerateOut, ReviewModerateIn, ReviewOut
from app.services import (
    admin_audit_service,
    admin_config_service,
    admin_master_service,
    admin_user_service,
    email_service,
)

router = APIRouter(prefix="/admin", tags=["admin"])


# ============================================================
# AD-02 USERS
# ============================================================

def _user_to_item(u) -> AdminUserListItem:
    item = AdminUserListItem.model_validate(u)
    item.roles = sorted(u.role_codes)
    return item


@router.get("/users", response_model=AdminUserListOut)
async def list_users(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    role: RoleCode | None = None,
    status_filter: UserStatus | None = Query(None, alias="status"),
    q: str | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
) -> AdminUserListOut:
    rows, total = await admin_user_service.list_users(
        db, user, role=role, status_filter=status_filter, q=q,
        offset=(page - 1) * size, limit=size,
    )
    return AdminUserListOut(
        items=[_user_to_item(u) for u in rows],
        total=total, page=page, size=size,
    )


@router.post("/users", response_model=AdminUserListItem, status_code=status.HTTP_201_CREATED)
async def create_user(
    data: AdminUserCreateIn,
    bg: BackgroundTasks,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AdminUserListItem:
    new_user = await admin_user_service.create_user(db, user, data)
    # Phase 1 step 4 (2026-05-06): gửi welcome email kèm temp password (fire-and-forget)
    bg.add_task(
        email_service.send_welcome,
        to_email=str(new_user.email),
        full_name=new_user.full_name,
        temp_password=data.password,
    )
    return _user_to_item(new_user)


@router.patch("/users/{user_id}", response_model=AdminUserListItem)
async def update_user(
    user_id: int,
    data: AdminUserUpdateIn,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AdminUserListItem:
    return _user_to_item(await admin_user_service.update_user(db, user, user_id, data))


@router.post("/users/{user_id}/lock", response_model=AdminUserListItem)
async def lock_user(
    user_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AdminUserListItem:
    return _user_to_item(await admin_user_service.lock_user(db, user, user_id))


@router.post("/users/{user_id}/unlock", response_model=AdminUserListItem)
async def unlock_user(
    user_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AdminUserListItem:
    return _user_to_item(await admin_user_service.unlock_user(db, user, user_id))


@router.post("/users/bulk-status", response_model=BulkUserStatusOut)
async def bulk_user_status(
    data: BulkUserStatusIn,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> BulkUserStatusOut:
    """AD-02 — khóa/mở/xóa HÀNG LOẠT user (vd: khóa 50 tài khoản spam 1 lần)."""
    affected, skipped = await admin_user_service.bulk_set_user_status(
        db, user, data.user_ids, data.action.value
    )
    return BulkUserStatusOut(affected_count=affected, skipped=skipped)


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def soft_delete_user(
    user_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    await admin_user_service.soft_delete_user(db, user, user_id)


@router.patch("/users/{user_id}/roles", response_model=AdminUserListItem)
async def replace_roles(
    user_id: int,
    data: AssignRolesIn,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AdminUserListItem:
    return _user_to_item(
        await admin_user_service.replace_roles(db, user, user_id, data.role_codes)
    )


@router.post("/students/import", response_model=BulkImportOut)
async def import_student_directory(
    data: BulkImportIn,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> BulkImportOut:
    return BulkImportOut(**await admin_user_service.import_student_directory(db, user, data))


# ============================================================
# AD-03 MASTER DATA
# ============================================================

@router.get("/faculties", response_model=list[FacultyOut])
async def list_faculties(db: Annotated[AsyncSession, Depends(get_db)]) -> list[FacultyOut]:
    return [FacultyOut.model_validate(f) for f in await admin_master_service.list_faculties(db)]


@router.post("/faculties", response_model=FacultyOut, status_code=status.HTTP_201_CREATED)
async def create_faculty(
    data: FacultyIn, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> FacultyOut:
    return FacultyOut.model_validate(
        await admin_master_service.create_faculty(db, user, data.faculty_code, data.faculty_name)
    )


@router.patch("/faculties/{faculty_id}", response_model=FacultyOut)
async def update_faculty(
    faculty_id: int, data: FacultyIn,
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> FacultyOut:
    return FacultyOut.model_validate(
        await admin_master_service.update_faculty(db, user, faculty_id, data.faculty_code, data.faculty_name)
    )


@router.delete("/faculties/{faculty_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_faculty(
    faculty_id: int, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    await admin_master_service.delete_faculty(db, user, faculty_id)


@router.get("/majors", response_model=list[MajorOut])
async def list_majors(
    db: Annotated[AsyncSession, Depends(get_db)],
    faculty_id: int | None = None,
) -> list[MajorOut]:
    return [MajorOut.model_validate(m) for m in await admin_master_service.list_majors(db, faculty_id)]


@router.post("/majors", response_model=MajorOut, status_code=status.HTTP_201_CREATED)
async def create_major(
    data: MajorIn, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> MajorOut:
    return MajorOut.model_validate(
        await admin_master_service.create_major(db, user, data.major_code, data.major_name, data.faculty_id)
    )


@router.patch("/majors/{major_id}", response_model=MajorOut)
async def update_major(
    major_id: int, data: MajorIn,
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> MajorOut:
    return MajorOut.model_validate(
        await admin_master_service.update_major(db, user, major_id, data.major_code, data.major_name, data.faculty_id)
    )


@router.delete("/majors/{major_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_major(
    major_id: int, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    await admin_master_service.delete_major(db, user, major_id)


@router.get("/classes", response_model=list[AcademicClassOut])
async def list_classes(
    db: Annotated[AsyncSession, Depends(get_db)],
    faculty_id: int | None = None, major_id: int | None = None,
) -> list[AcademicClassOut]:
    items = await admin_master_service.list_classes(db, faculty_id, major_id)
    return [AcademicClassOut.model_validate(c) for c in items]


@router.post("/classes", response_model=AcademicClassOut, status_code=status.HTTP_201_CREATED)
async def create_class(
    data: AcademicClassIn, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> AcademicClassOut:
    return AcademicClassOut.model_validate(
        await admin_master_service.create_class(
            db, user, data.class_code, data.class_name, data.faculty_id, data.major_id
        )
    )


@router.patch("/classes/{class_id}", response_model=AcademicClassOut)
async def update_class(
    class_id: int, data: AcademicClassIn,
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> AcademicClassOut:
    return AcademicClassOut.model_validate(
        await admin_master_service.update_class(
            db, user, class_id, data.class_code, data.class_name, data.faculty_id, data.major_id
        )
    )


@router.delete("/classes/{class_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_class(
    class_id: int, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    await admin_master_service.delete_class(db, user, class_id)


# ============================================================
# AD-04 SYSTEM CONFIGS
# ============================================================

@router.get("/configs", response_model=list[ConfigOut])
async def list_configs(
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> list[ConfigOut]:
    return [ConfigOut(**c) for c in await admin_config_service.list_configs(db, user)]


@router.get("/configs/{key}", response_model=ConfigOut)
async def get_config(
    key: str, user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> ConfigOut:
    return ConfigOut(**await admin_config_service.get_config(db, user, key))


@router.patch("/configs/{key}", response_model=ConfigOut)
async def update_config(
    key: str, data: ConfigUpdateIn,
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> ConfigOut:
    return ConfigOut(**await admin_config_service.update_config(db, user, key, data.config_value))


# ============================================================
# AD-06 AUDIT LOGS + REVIEW MODERATION
# ============================================================

@router.get("/audit-logs", response_model=AuditLogListOut)
async def list_audit_logs(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    user_id: int | None = None,
    action_type: str | None = None,
    entity_name: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=500),
) -> AuditLogListOut:
    rows, total = await admin_audit_service.list_audit_logs(
        db, user,
        user_id=user_id, action_type=action_type, entity_name=entity_name,
        date_from=date_from, date_to=date_to,
        offset=(page - 1) * size, limit=size,
    )
    items = []
    for r in rows:
        # Build dict thủ công vì IPv4Address từ asyncpg INET không cast tự động
        # khi dùng model_validate(from_attributes=True) — Pydantic v2 strict.
        items.append(AuditLogOut(
            log_id=r.log_id,
            user_id=r.user_id,
            action_type=r.action_type,
            entity_name=r.entity_name,
            entity_id=r.entity_id,
            ip_address=str(r.ip_address) if r.ip_address is not None else None,
            details_json=r.details_json,
            created_at=r.created_at,
        ))
    return AuditLogListOut(items=items, total=total, page=page, size=size)


@router.get("/reviews", response_model=list[ReviewOut])
async def list_all_reviews(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    contest_id: int | None = None,
    only_hidden: bool = False,
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
) -> list[ReviewOut]:
    rows, _ = await admin_audit_service.list_all_reviews(
        db, user, contest_id, only_hidden, (page - 1) * size, size
    )
    return [ReviewOut.model_validate(r) for r in rows]


@router.patch("/reviews/{review_id}/moderate", response_model=ReviewOut)
async def moderate_review(
    review_id: int, data: ReviewModerateIn,
    user: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)],
) -> ReviewOut:
    review = await admin_audit_service.moderate_review(
        db, user, review_id, data.is_visible, data.moderation_note
    )
    return ReviewOut.model_validate(review)


@router.patch("/reviews/bulk-moderate", response_model=BulkModerateOut)
async def bulk_moderate_reviews(
    data: BulkModerateIn,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> BulkModerateOut:
    """AD-06 — ẩn/hiện HÀNG LOẠT review (vd: ẩn 50+ review spam 1 lần)."""
    count, not_found = await admin_audit_service.bulk_moderate_reviews(
        db, user, data.review_ids, data.is_visible, data.moderation_note
    )
    return BulkModerateOut(moderated_count=count, not_found=not_found)


# ============================================================
# Sprint 21 hotfix (2026-05-09): backfill host_faculty_id
# Dev/maintenance utility: contests cũ tạo trước khi BE auto-inject
# faculty từ organizer profile có host_faculty_id = NULL → BCN approval
# queue filter loại bỏ → BCN không thấy duyệt được. Endpoint này chạy
# SQL UPDATE để fix các contests legacy.
# ============================================================

@router.post("/backfill-host-faculty")
async def backfill_host_faculty(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Admin maintenance — backfill host_faculty_id cho contests đang NULL.

    Lấy faculty_id từ profile Organizer của user tạo contest (created_by).
    Contest nào creator không có Organizer profile (vd admin tạo) sẽ skip.
    """
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN")

    from app.models.contest import Contest
    from app.models.master_data import Organizer

    # Tìm contests có host_faculty_id NULL
    null_stmt = select(Contest).where(Contest.host_faculty_id.is_(None))
    contests = (await db.execute(null_stmt)).scalars().all()

    updated: list[dict] = []
    skipped: list[dict] = []
    for c in contests:
        org_stmt = select(Organizer).where(Organizer.user_id == c.created_by)
        org = (await db.execute(org_stmt)).scalar_one_or_none()
        if org is not None and org.faculty_id is not None:
            c.host_faculty_id = org.faculty_id
            updated.append({
                "contest_id": c.contest_id,
                "slug": c.slug,
                "title": c.title,
                "faculty_id": org.faculty_id,
            })
        else:
            skipped.append({
                "contest_id": c.contest_id,
                "slug": c.slug,
                "reason": "creator chưa có Organizer profile hoặc không có faculty_id",
            })

    await db.commit()
    return {
        "updated_count": len(updated),
        "skipped_count": len(skipped),
        "updated": updated,
        "skipped": skipped,
    }


# ============================================================
# Sprint 25 P2-C1 (2026-05-09): Faculty cert templates CRUD
# Quản lý mẫu chứng nhận theo khoa — BCN/HOD scope.
# Admin xem/sửa được tất cả khoa (qua query param), HOD chỉ khoa của mình.
# ============================================================

@router.get("/faculty-cert-templates")
async def list_faculty_cert_templates(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list:
    """List templates của faculty user (HOD scope) hoặc tất cả (Admin)."""
    from app.models.certificate import FacultyCertTemplate
    from app.models.master_data import DepartmentHead
    from app.schemas.certificate import FacultyCertTemplateOut

    is_admin = "ADMIN" in user.role_codes
    is_hod = "HOD" in user.role_codes
    if not (is_admin or is_hod):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "Cần role ADMIN hoặc HOD")

    stmt = select(FacultyCertTemplate)
    if is_hod and not is_admin:
        dh = (await db.execute(
            select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
        )).scalar_one_or_none()
        if dh is None:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN, "User chưa có profile HOD")
        stmt = stmt.where(FacultyCertTemplate.faculty_id == dh.faculty_id)

    rows = (await db.execute(stmt.order_by(FacultyCertTemplate.template_id.desc()))).scalars().all()
    return [FacultyCertTemplateOut.model_validate(r).model_dump(mode='json') for r in rows]


@router.post("/faculty-cert-templates", status_code=status.HTTP_201_CREATED)
async def create_faculty_cert_template(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    data: Annotated[dict, Body()],
) -> dict:
    """HOD tạo template mới cho khoa của mình. Admin truyền faculty_id qua body."""
    from app.models.certificate import FacultyCertTemplate
    from app.models.master_data import DepartmentHead
    from app.schemas.certificate import FacultyCertTemplateOut

    is_admin = "ADMIN" in user.role_codes
    is_hod = "HOD" in user.role_codes
    if not (is_admin or is_hod):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "Cần role ADMIN hoặc HOD")

    # Auto inject faculty_id từ HOD profile (nếu không phải admin truyền tay)
    if is_admin and "faculty_id" in data:
        faculty_id = data["faculty_id"]
    else:
        dh = (await db.execute(
            select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
        )).scalar_one_or_none()
        if dh is None:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN, "User chưa có profile HOD")
        faculty_id = dh.faculty_id

    tpl = FacultyCertTemplate(
        faculty_id=faculty_id,
        name=data.get("name") or "",
        layout_description=data.get("layout_description") or "",
        signers=data.get("signers") or "",
        is_active=bool(data.get("is_active", False)),
        created_by=user.user_id,
    )
    db.add(tpl)
    await db.commit()
    await db.refresh(tpl)
    return FacultyCertTemplateOut.model_validate(tpl).model_dump(mode='json')


@router.patch("/faculty-cert-templates/{template_id}")
async def update_faculty_cert_template(
    template_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    data: Annotated[dict, Body()],
) -> dict:
    """HOD update template trong khoa mình. Admin update bất kỳ."""
    from app.models.certificate import FacultyCertTemplate
    from app.models.master_data import DepartmentHead
    from app.schemas.certificate import FacultyCertTemplateOut

    tpl = await db.get(FacultyCertTemplate, template_id)
    if tpl is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Template not found")

    is_admin = "ADMIN" in user.role_codes
    is_hod = "HOD" in user.role_codes
    if not (is_admin or is_hod):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN hoặc HOD")

    if is_hod and not is_admin:
        dh = (await db.execute(
            select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
        )).scalar_one_or_none()
        if dh is None or dh.faculty_id != tpl.faculty_id:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "HOD chỉ sửa template của khoa mình")

    for key in ("name", "layout_description", "signers", "is_active"):
        if key in data and data[key] is not None:
            setattr(tpl, key, data[key])
    await db.commit()
    await db.refresh(tpl)
    return FacultyCertTemplateOut.model_validate(tpl).model_dump(mode='json')


@router.delete("/faculty-cert-templates/{template_id}",
               status_code=status.HTTP_204_NO_CONTENT)
async def delete_faculty_cert_template(
    template_id: int,
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """HOD xóa template trong khoa mình. Admin xóa bất kỳ."""
    from app.models.certificate import FacultyCertTemplate
    from app.models.master_data import DepartmentHead

    tpl = await db.get(FacultyCertTemplate, template_id)
    if tpl is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Template not found")

    is_admin = "ADMIN" in user.role_codes
    is_hod = "HOD" in user.role_codes
    if not (is_admin or is_hod):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN hoặc HOD")

    if is_hod and not is_admin:
        dh = (await db.execute(
            select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
        )).scalar_one_or_none()
        if dh is None or dh.faculty_id != tpl.faculty_id:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "HOD chỉ xóa template của khoa mình")

    await db.delete(tpl)
    await db.commit()
    return None


# ============================================================
# AD-04 BACKUP — pg_dump schema ptit_contest (2026-06-29)
# FE (admin_shell._BackupRestoreScreen) gọi POST /admin/backup. Trước đây
# endpoint chưa tồn tại → 404 → FE hiện "đã xóa hoặc di chuyển". Image backend
# đã có postgresql-client (pg_dump). Restore vẫn để DBA chạy CLI thủ công.
# ============================================================

@router.post("/backup")
async def create_backup(user: CurrentUser) -> dict:
    """Admin tạo backup pg_dump schema ptit_contest → file .sql trong /backups."""
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN")

    import asyncio
    import os
    from urllib.parse import urlparse, unquote

    from app.config import settings

    # database_url dạng SQLAlchemy async: postgresql+asyncpg://user:pass@host:port/db
    raw = settings.database_url.replace("+asyncpg", "").replace("+psycopg", "")
    p = urlparse(raw)
    db_user = unquote(p.username or "ptit_contest")
    db_pass = unquote(p.password or "")
    db_host = p.hostname or "localhost"
    db_port = str(p.port or 5432)
    db_name = (p.path or "/ptit_contest_db").lstrip("/")

    backup_dir = os.environ.get("BACKUP_DIR", "/backups")
    try:
        os.makedirs(backup_dir, exist_ok=True)
    except OSError:
        backup_dir = "/tmp/backups"
        os.makedirs(backup_dir, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ptit_contest_{ts}.sql"
    filepath = os.path.join(backup_dir, filename)

    cmd = [
        "pg_dump",
        "-h", db_host,
        "-p", db_port,
        "-U", db_user,
        "-d", db_name,
        "--schema=ptit_contest",
        "--no-owner",
        "--no-privileges",
        "-f", filepath,
    ]
    env = {**os.environ, "PGPASSWORD": db_pass}

    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            env=env,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        _, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
    except FileNotFoundError:
        raise HTTPException(
            status.HTTP_501_NOT_IMPLEMENTED,
            "pg_dump không có trong container. Cần postgresql-client trong image.",
        )
    except asyncio.TimeoutError:
        raise HTTPException(status.HTTP_504_GATEWAY_TIMEOUT, "pg_dump timeout (>120s)")

    if proc.returncode != 0:
        msg = stderr.decode(errors="replace")[:500] if stderr else "unknown error"
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR,
                            f"pg_dump lỗi: {msg}")

    size_bytes = os.path.getsize(filepath)
    return {
        "filename": filename,
        "path": filepath,
        "size_mb": round(size_bytes / (1024 * 1024), 3),
        "created_at": datetime.now().isoformat(),
    }


@router.get("/backup/download/{filename}")
async def download_backup(filename: str, user: CurrentUser) -> FileResponse:
    """Admin tải file backup .sql đã tạo từ thư mục BACKUP_DIR."""
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN")

    import os

    # Chặn path traversal: chỉ cho phép basename, đuôi .sql.
    safe = os.path.basename(filename)
    if safe != filename or not safe.endswith(".sql"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tên file không hợp lệ")

    backup_dir = os.environ.get("BACKUP_DIR", "/backups")
    filepath = os.path.join(backup_dir, safe)
    if not os.path.isfile(filepath):
        # thử fallback /tmp/backups (khi /backups không ghi được lúc tạo)
        filepath = os.path.join("/tmp/backups", safe)
        if not os.path.isfile(filepath):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "File backup không tồn tại")

    return FileResponse(
        filepath,
        media_type="application/sql",
        filename=safe,
    )


# ============================================================
# AD-05 EXPORT EXCEL — danh sách user + audit log (2026-06-29)
# Reuse openpyxl. FE gọi qua exportXlsxFromEndpoint (responseType bytes).
# ============================================================

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_xlsx(sheet_title: str, headers: list[str], rows: list[list]) -> bytes:
    """Dựng workbook 1 sheet: header bold + nền xám, auto width, trả bytes."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel giới hạn 31 ký tự tên sheet

    head_fill = PatternFill("solid", fgColor="E5283C")  # ptitRed
    head_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # auto width thô theo độ dài nội dung
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in rows:
            if col - 1 < len(row) and row[col - 1] is not None:
                max_len = max(max_len, len(str(row[col - 1])))
        ws.column_dimensions[letter].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_stream(data: bytes, filename: str) -> StreamingResponse:
    def _iter():
        chunk = 64 * 1024
        for i in range(0, len(data), chunk):
            yield data[i:i + chunk]

    return StreamingResponse(
        _iter(),
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )


@router.get("/users/export.xlsx")
async def export_users_xlsx(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    role: RoleCode | None = None,
    status_filter: UserStatus | None = Query(None, alias="status"),
    q: str | None = None,
) -> StreamingResponse:
    """Admin xuất danh sách user ra Excel (theo bộ lọc role/status/q hiện tại)."""
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN")

    rows, _ = await admin_user_service.list_users(
        db, user, role=role, status_filter=status_filter, q=q,
        offset=0, limit=10000,
    )
    headers = ["ID", "Email", "Họ tên", "SĐT", "Roles", "Trạng thái",
               "Đăng nhập gần nhất", "Ngày tạo"]
    data_rows = []
    for u in rows:
        item = _user_to_item(u)
        data_rows.append([
            item.user_id,
            item.email,
            item.full_name,
            item.phone or "",
            ", ".join(item.roles),
            item.status.value if hasattr(item.status, "value") else str(item.status),
            item.last_login_at.strftime("%d/%m/%Y %H:%M") if item.last_login_at else "—",
            item.created_at.strftime("%d/%m/%Y %H:%M") if item.created_at else "",
        ])

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = _build_xlsx("Danh sách user", headers, data_rows)
    return _xlsx_stream(xlsx, f"danh-sach-user_{ts}.xlsx")


@router.get("/audit-logs/export.xlsx")
async def export_audit_logs_xlsx(
    user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    user_id: int | None = None,
    action_type: str | None = None,
    entity_name: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
) -> StreamingResponse:
    """Admin xuất audit log ra Excel (theo bộ lọc hiện tại, tối đa 10000 dòng)."""
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần role ADMIN")

    rows, _ = await admin_audit_service.list_audit_logs(
        db, user,
        user_id=user_id, action_type=action_type, entity_name=entity_name,
        date_from=date_from, date_to=date_to,
        offset=0, limit=10000,
    )
    headers = ["Log ID", "Thời gian", "User ID", "Hành động", "Đối tượng",
               "Entity ID", "IP", "Chi tiết"]
    data_rows = []
    for r in rows:
        details = r.details_json
        data_rows.append([
            r.log_id,
            r.created_at.strftime("%d/%m/%Y %H:%M:%S") if r.created_at else "",
            r.user_id if r.user_id is not None else "",
            r.action_type,
            r.entity_name,
            r.entity_id or "",
            str(r.ip_address) if r.ip_address is not None else "",
            str(details) if details else "",
        ])

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = _build_xlsx("Audit log", headers, data_rows)
    return _xlsx_stream(xlsx, f"audit-log_{ts}.xlsx")
