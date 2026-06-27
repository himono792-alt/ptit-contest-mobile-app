"""Reports / aggregate stats (GV-07, BCN-03, BCN-05, AD-05)."""

from datetime import datetime, timedelta as _td, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.certificate import IssuedCertificate
from app.models.contest import Contest, ContestRound
from app.models.entry import ContestEntry, Team, TeamMember
from app.models.enums import (
    ApprovalStatus,
    ContestStatus,
    EntryType,
    RegistrationStatus,
    RoleCode,
    SubmissionStatus,
    UserStatus,
)
from app.models.identity import AppUser, Role, UserRole
from app.models.judging import ContestResult, RoundResult
from app.models.master_data import (
    DepartmentHead,
    Faculty,
    Student,
    StudentDirectory,
)
from app.models.review import ContestReview
from app.models.submission import Submission
from app.models.workflow import WorkflowApproval


# ---------- Helpers ----------

# PTIT brand styles cho xlsx export (dùng chung 2 hàm export bên dưới)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_headers(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize(ws, max_width: int = 50) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


async def _ensure_organizer_or_admin(user, contest):
    if "ADMIN" in user.role_codes:
        return
    if "ORGANIZER" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ORGANIZER hoặc ADMIN")
    if contest.created_by != user.user_id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Không phải BTC contest này")


async def _ensure_hod_or_admin(db: AsyncSession, user) -> int | None:
    """Trả faculty_id để filter (None = ADMIN xem tất)."""
    if "ADMIN" in user.role_codes:
        return None
    if "HOD" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần HOD hoặc ADMIN")
    dh = (await db.execute(
        select(DepartmentHead).where(DepartmentHead.user_id == user.user_id)
    )).scalar_one_or_none()
    if dh is None:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "User chưa có profile HOD")
    return dh.faculty_id


# ============================================================
# GV-07 contest stats
# ============================================================

async def contest_stats(db: AsyncSession, user, contest_id: int) -> dict:
    contest = await db.get(Contest, contest_id)
    if contest is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contest not found")
    await _ensure_organizer_or_admin(user, contest)

    # Entries by status
    entries_by_status = (await db.execute(
        select(
            ContestEntry.registration_status,
            func.count().label("c"),
        )
        .where(ContestEntry.contest_id == contest_id)
        .group_by(ContestEntry.registration_status)
    )).all()
    by_status = {row.registration_status: row.c for row in entries_by_status}
    total_entries = sum(by_status.values())

    # Submissions
    sub_stats = (await db.execute(
        select(
            func.count().label("total"),
            func.sum(case((Submission.status == SubmissionStatus.SUBMITTED, 1), else_=0)).label("submitted"),
            func.sum(case((Submission.status == SubmissionStatus.LATE, 1), else_=0)).label("late"),
        )
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .where(ContestRound.contest_id == contest_id)
    )).one()

    # Rounds
    rounds_count = (await db.execute(
        select(func.count()).select_from(ContestRound).where(ContestRound.contest_id == contest_id)
    )).scalar_one()

    rounds_with_results = (await db.execute(
        select(func.count(func.distinct(RoundResult.round_id)))
        .select_from(RoundResult)
        .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
        .where(ContestRound.contest_id == contest_id)
    )).scalar_one()

    # Final score stats (avg + pass rate top 50%)
    final_stats = (await db.execute(
        select(
            func.avg(ContestResult.final_score).label("avg_score"),
            func.count().label("c"),
        )
        .where(
            ContestResult.contest_id == contest_id,
            ContestResult.final_score.is_not(None),
        )
    )).one()
    avg_score = final_stats.avg_score
    total_results = final_stats.c

    pass_rate = None
    if total_results > 0:
        # Top 50% = rank <= total/2
        top_half = (await db.execute(
            select(func.count())
            .select_from(ContestResult)
            .where(
                ContestResult.contest_id == contest_id,
                ContestResult.rank_no.is_not(None),
                ContestResult.rank_no <= max(1, total_results // 2),
            )
        )).scalar_one()
        pass_rate = Decimal(top_half * 100) / Decimal(total_results)

    # Reviews
    rev_stats = (await db.execute(
        select(
            func.avg(ContestReview.rating).label("avg"),
            func.count().label("c"),
        )
        .where(
            ContestReview.contest_id == contest_id,
            ContestReview.is_visible.is_(True),
        )
    )).one()

    return {
        "contest_id": contest_id,
        "contest_title": contest.title,
        "contest_status": contest.status.value,
        "total_entries": total_entries,
        "approved_entries": by_status.get(RegistrationStatus.APPROVED, 0),
        "pending_entries": by_status.get(RegistrationStatus.PENDING, 0),
        "rejected_entries": by_status.get(RegistrationStatus.REJECTED, 0),
        "cancelled_entries": by_status.get(RegistrationStatus.CANCELLED, 0),
        "total_submissions": sub_stats.total or 0,
        "submitted_count": sub_stats.submitted or 0,
        "late_count": sub_stats.late or 0,
        "rounds_count": rounds_count,
        "rounds_with_results": rounds_with_results,
        "average_final_score": avg_score,
        "pass_rate_percent": pass_rate,
        "avg_review_rating": rev_stats.avg,
        "review_count": rev_stats.c,
    }


# ============================================================
# BCN-03 monitor
# ============================================================

async def monitor_contests(
    db: AsyncSession, user, faculty_filter: int | None = None
) -> list[dict]:
    user_faculty = await _ensure_hod_or_admin(db, user)
    if user_faculty is not None:
        faculty_filter = user_faculty  # HOD chỉ xem khoa mình

    base = select(Contest).order_by(Contest.start_at.desc())
    if faculty_filter is not None:
        base = base.where(Contest.host_faculty_id == faculty_filter)
    contests = (await db.execute(base)).scalars().all()

    items: list[dict] = []
    for c in contests:
        # Count entries
        entries_total = (await db.execute(
            select(func.count())
            .select_from(ContestEntry)
            .where(ContestEntry.contest_id == c.contest_id)
        )).scalar_one()
        entries_approved = (await db.execute(
            select(func.count())
            .select_from(ContestEntry)
            .where(
                ContestEntry.contest_id == c.contest_id,
                ContestEntry.registration_status == RegistrationStatus.APPROVED,
            )
        )).scalar_one()

        # Submissions
        sub_count = (await db.execute(
            select(func.count())
            .select_from(Submission)
            .join(ContestRound, ContestRound.round_id == Submission.round_id)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()

        # Rounds
        rounds_total = (await db.execute(
            select(func.count())
            .select_from(ContestRound)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()
        rounds_done = (await db.execute(
            select(func.count(func.distinct(RoundResult.round_id)))
            .select_from(RoundResult)
            .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
            .where(ContestRound.contest_id == c.contest_id)
        )).scalar_one()

        reg_pct = None
        if c.max_entries:
            reg_pct = round(entries_total / c.max_entries * 100, 1)

        sub_pct = None
        if entries_approved > 0:
            sub_pct = round(sub_count / entries_approved * 100, 1)

        judg_pct = None
        if rounds_total > 0:
            judg_pct = round(rounds_done / rounds_total * 100, 1)

        items.append({
            "contest_id": c.contest_id,
            "title": c.title,
            "slug": c.slug,
            "status": c.status.value,
            "start_at": c.start_at.isoformat() if c.start_at else None,
            "end_at": c.end_at.isoformat() if c.end_at else None,
            "registration_pct": reg_pct,
            "submission_pct": sub_pct,
            "judging_pct": judg_pct,
            "total_entries": entries_total,
            "total_submissions": sub_count,
        })
    return items


# ============================================================
# BCN-05 faculty summary
# ============================================================

async def faculty_summary(
    db: AsyncSession, user, faculty_id: int | None, year: int
) -> dict:
    user_faculty = await _ensure_hod_or_admin(db, user)
    target_faculty = user_faculty if user_faculty is not None else faculty_id
    if target_faculty is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Admin cần truyền faculty_id query param",
        )

    faculty = await db.get(Faculty, target_faculty)
    if faculty is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Faculty not found")

    # Year boundary
    year_start = datetime(year, 1, 1, tzinfo=timezone.utc)
    year_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    # Contests of faculty in year
    base_filter = [
        Contest.host_faculty_id == target_faculty,
        Contest.created_at >= year_start,
        Contest.created_at < year_end,
    ]

    total_contests = (await db.execute(
        select(func.count()).select_from(Contest).where(*base_filter)
    )).scalar_one()
    finished = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status == ContestStatus.FINISHED)
    )).scalar_one()
    ongoing = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status.in_([ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING]))
    )).scalar_one()
    draft_pending = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*base_filter, Contest.status.in_([ContestStatus.DRAFT, ContestStatus.PROPOSED, ContestStatus.REVISION_REQUESTED]))
    )).scalar_one()

    # Entries + unique students in faculty contests
    entries_total = (await db.execute(
        select(func.count())
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*base_filter)
    )).scalar_one()
    unique_students = (await db.execute(
        select(func.count(func.distinct(ContestEntry.student_id)))
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*base_filter, ContestEntry.student_id.is_not(None))
    )).scalar_one()

    # Awards
    awards_count = (await db.execute(
        select(func.count())
        .select_from(ContestResult)
        .join(Contest, Contest.contest_id == ContestResult.contest_id)
        .where(*base_filter, ContestResult.award_title.is_not(None))
    )).scalar_one()

    # Avg rating
    avg_rating = (await db.execute(
        select(func.avg(ContestReview.rating))
        .select_from(ContestReview)
        .join(Contest, Contest.contest_id == ContestReview.contest_id)
        .where(*base_filter, ContestReview.is_visible.is_(True))
    )).scalar_one()

    return {
        "faculty_id": target_faculty,
        "faculty_code": faculty.faculty_code,
        "faculty_name": faculty.faculty_name,
        "year": year,
        "total_contests": total_contests,
        "contests_finished": finished,
        "contests_ongoing": ongoing,
        "contests_draft_or_pending": draft_pending,
        "total_unique_students": unique_students,
        "total_entries": entries_total,
        "total_awards": awards_count,
        "avg_completion_rate": (
            Decimal(finished * 100) / Decimal(total_contests)
            if total_contests > 0 else None
        ),
        "avg_rating": avg_rating,
    }


# ============================================================
# AD-05 system summary
# ============================================================

async def system_summary(db: AsyncSession, user, year: int) -> dict:
    if "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ADMIN")

    year_start = datetime(year, 1, 1, tzinfo=timezone.utc)
    year_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    # User counts
    total_users = (await db.execute(
        select(func.count()).select_from(AppUser)
        .where(AppUser.status != UserStatus.DELETED)
    )).scalar_one()

    counts_by_role = {}
    for rc in [RoleCode.STUDENT, RoleCode.ORGANIZER, RoleCode.JUDGE, RoleCode.HOD, RoleCode.ADMIN]:
        cnt = (await db.execute(
            select(func.count(func.distinct(AppUser.user_id)))
            .select_from(AppUser)
            .join(UserRole, UserRole.user_id == AppUser.user_id)
            .join(Role, Role.role_id == UserRole.role_id)
            .where(
                AppUser.status == UserStatus.ACTIVE,
                Role.role_code == rc,
            )
        )).scalar_one()
        counts_by_role[rc] = cnt

    # Contests
    contests_in_year = [Contest.created_at >= year_start, Contest.created_at < year_end]
    total_contests = (await db.execute(
        select(func.count()).select_from(Contest).where(*contests_in_year)
    )).scalar_one()
    published_or_after = (await db.execute(
        select(func.count()).select_from(Contest).where(
            *contests_in_year,
            Contest.status.in_([
                ContestStatus.PUBLISHED, ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED,
                ContestStatus.ONGOING, ContestStatus.FINISHED,
            ]),
        )
    )).scalar_one()
    finished = (await db.execute(
        select(func.count()).select_from(Contest)
        .where(*contests_in_year, Contest.status == ContestStatus.FINISHED)
    )).scalar_one()

    # Entries / submissions
    entries_total = (await db.execute(
        select(func.count())
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(*contests_in_year)
    )).scalar_one()
    submissions_total = (await db.execute(
        select(func.count())
        .select_from(Submission)
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .join(Contest, Contest.contest_id == ContestRound.contest_id)
        .where(*contests_in_year)
    )).scalar_one()

    # Certificates
    certs_total = (await db.execute(
        select(func.count())
        .select_from(IssuedCertificate)
        .join(ContestResult, ContestResult.contest_result_id == IssuedCertificate.contest_result_id)
        .join(Contest, Contest.contest_id == ContestResult.contest_id)
        .where(*contests_in_year, IssuedCertificate.revoked_at.is_(None))
    )).scalar_one()

    # Reviews
    rev_stats = (await db.execute(
        select(
            func.count().label("c"),
            func.avg(ContestReview.rating).label("avg"),
        )
        .select_from(ContestReview)
        .join(Contest, Contest.contest_id == ContestReview.contest_id)
        .where(*contests_in_year, ContestReview.is_visible.is_(True))
    )).one()

    return {
        "year": year,
        "total_users": total_users,
        "students_active": counts_by_role[RoleCode.STUDENT],
        "organizers": counts_by_role[RoleCode.ORGANIZER],
        "judges": counts_by_role[RoleCode.JUDGE],
        "department_heads": counts_by_role[RoleCode.HOD],
        "admins": counts_by_role[RoleCode.ADMIN],
        "total_contests": total_contests,
        "contests_published_or_after": published_or_after,
        "contests_finished": finished,
        "total_entries": entries_total,
        "total_submissions": submissions_total,
        "total_certificates_issued": certs_total,
        "total_reviews": rev_stats.c or 0,
        "avg_review_rating": rev_stats.avg,
    }


# ============================================================
# Phase 2 sprint 1 step 3 (2026-05-06): Excel export contest results
# ============================================================

async def export_contest_results_xlsx(
    db: AsyncSession, user: AppUser, contest_id: int
) -> tuple[bytes, str]:
    """Render contest results to xlsx (4 sheets).

    Trả tuple (bytes_data, suggested_filename).
    Caller wrap vào StreamingResponse với content-type
    application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.

    Sheets:
    - Tổng quan: rank, SV/team, mã SV, total_score, award
    - Vòng: rank theo round, total_score, average_score, is_passed
    - Submissions: round_no, entry, status, submitted_at
    - Metadata: contest info, generated by, generated at

    Permission: BTC của contest, BCN, ADMIN.
    """
    contest = await db.get(Contest, contest_id)
    if contest is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contest not found")
    await _ensure_organizer_or_admin(user, contest)

    # ----- Fetch data -----
    # Sheet 1: Tổng quan — JOIN ContestResult → ContestEntry → Student/Team → AppUser
    overall_stmt = (
        select(ContestResult, ContestEntry)
        .join(ContestEntry, ContestEntry.entry_id == ContestResult.entry_id)
        .where(ContestResult.contest_id == contest_id)
        .order_by(ContestResult.rank_no.nulls_last())
    )
    overall_rows = (await db.execute(overall_stmt)).all()

    # Resolve display name + code cho từng entry
    async def _entry_display(entry: ContestEntry) -> tuple[str, str]:
        """Trả (display_name, identifier_code). Cho INDIVIDUAL: SV name + MSSV.
        Cho TEAM: team_name + team_id."""
        if entry.entry_type == EntryType.INDIVIDUAL and entry.student_id:
            stmt = (
                select(Student.student_code, AppUser.full_name)
                .join(AppUser, AppUser.user_id == Student.user_id)
                .where(Student.student_id == entry.student_id)
            )
            row = (await db.execute(stmt)).first()
            if row:
                return row.full_name, row.student_code
            return "(SV không xác định)", str(entry.student_id)
        if entry.entry_type == EntryType.TEAM and entry.team_id:
            team = await db.get(Team, entry.team_id)
            if team:
                return team.team_name, f"Team #{team.team_id}"
            return "(Team không xác định)", f"Team #{entry.team_id}"
        return "(Không xác định)", str(entry.entry_id)

    # Sheet 2: ContestRound + RoundResult
    rounds_stmt = (
        select(ContestRound)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no)
    )
    rounds = list((await db.execute(rounds_stmt)).scalars().all())

    round_results_stmt = (
        select(RoundResult, ContestRound)
        .join(ContestRound, ContestRound.round_id == RoundResult.round_id)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no, RoundResult.rank_no.nulls_last())
    )
    round_result_rows = (await db.execute(round_results_stmt)).all()

    # Sheet 3: Submissions
    submissions_stmt = (
        select(Submission, ContestRound)
        .join(ContestRound, ContestRound.round_id == Submission.round_id)
        .where(ContestRound.contest_id == contest_id)
        .order_by(ContestRound.round_no, Submission.entry_id)
    )
    submission_rows = (await db.execute(submissions_stmt)).all()

    # ----- Build workbook -----
    wb = Workbook()
    # openpyxl tạo sẵn 1 sheet "Sheet" — em rename thành "Tổng quan"
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws2 = wb.create_sheet("Vòng - kết quả")
    ws3 = wb.create_sheet("Submissions")
    ws4 = wb.create_sheet("Metadata")

    # ----- Sheet 1: Tổng quan -----
    _write_headers(ws1, ["Hạng", "Tên / Team", "Mã SV / Team ID", "Tổng điểm", "Giải thưởng", "BCN duyệt"])
    for ridx, (cr, entry) in enumerate(overall_rows, start=2):
        name, code = await _entry_display(entry)
        ws1.cell(row=ridx, column=1, value=cr.rank_no or "—")
        ws1.cell(row=ridx, column=2, value=name)
        ws1.cell(row=ridx, column=3, value=code)
        score_cell = ws1.cell(row=ridx, column=4,
                               value=float(cr.final_score) if cr.final_score is not None else None)
        score_cell.number_format = "0.00"
        ws1.cell(row=ridx, column=5, value=cr.award_title or "")
        ws1.cell(row=ridx, column=6, value=cr.bcn_approval_status.value if cr.bcn_approval_status else "")
    _autosize(ws1)

    # ----- Sheet 2: Vòng - kết quả -----
    _write_headers(ws2, ["Vòng", "Tên vòng", "Hạng", "Entry ID", "Tên / Team", "Tổng điểm", "Điểm TB", "Đậu?"])
    # Cache entry display để không query lặp
    entry_display_cache: dict[int, tuple[str, str]] = {}
    for cr, _entry in overall_rows:
        entry_display_cache[_entry.entry_id] = await _entry_display(_entry)

    row_idx = 2
    for rr, rnd in round_result_rows:
        # Lookup entry display (có thể không có trong overall_rows nếu entry bị eliminate giữa chừng)
        if rr.entry_id not in entry_display_cache:
            ent = await db.get(ContestEntry, rr.entry_id)
            if ent is not None:
                entry_display_cache[rr.entry_id] = await _entry_display(ent)
            else:
                entry_display_cache[rr.entry_id] = ("(deleted)", str(rr.entry_id))
        name, _code = entry_display_cache[rr.entry_id]
        ws2.cell(row=row_idx, column=1, value=rnd.round_no)
        ws2.cell(row=row_idx, column=2, value=rnd.round_name)
        ws2.cell(row=row_idx, column=3, value=rr.rank_no or "—")
        ws2.cell(row=row_idx, column=4, value=rr.entry_id)
        ws2.cell(row=row_idx, column=5, value=name)
        c1 = ws2.cell(row=row_idx, column=6, value=float(rr.total_score) if rr.total_score else None)
        c1.number_format = "0.00"
        c2 = ws2.cell(row=row_idx, column=7, value=float(rr.average_score) if rr.average_score else None)
        c2.number_format = "0.00"
        ws2.cell(row=row_idx, column=8, value="✓" if rr.is_passed else ("✗" if rr.is_passed is False else ""))
        row_idx += 1
    _autosize(ws2)

    # ----- Sheet 3: Submissions -----
    _write_headers(ws3, ["Vòng", "Tên vòng", "Submission ID", "Entry ID", "Tên / Team",
                         "Trạng thái", "Phiên bản", "Đã nộp lúc"])
    row_idx = 2
    for sub, rnd in submission_rows:
        if sub.entry_id not in entry_display_cache:
            ent = await db.get(ContestEntry, sub.entry_id)
            if ent is not None:
                entry_display_cache[sub.entry_id] = await _entry_display(ent)
            else:
                entry_display_cache[sub.entry_id] = ("(deleted)", str(sub.entry_id))
        name, _code = entry_display_cache[sub.entry_id]
        ws3.cell(row=row_idx, column=1, value=rnd.round_no)
        ws3.cell(row=row_idx, column=2, value=rnd.round_name)
        ws3.cell(row=row_idx, column=3, value=sub.submission_id)
        ws3.cell(row=row_idx, column=4, value=sub.entry_id)
        ws3.cell(row=row_idx, column=5, value=name)
        ws3.cell(row=row_idx, column=6, value=sub.status.value if sub.status else "")
        ws3.cell(row=row_idx, column=7, value=sub.current_version_no)
        ws3.cell(row=row_idx, column=8,
                 value=sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "")
        row_idx += 1
    _autosize(ws3)

    # ----- Sheet 4: Metadata -----
    _write_headers(ws4, ["Trường thông tin", "Giá trị"])
    metadata_rows = [
        ("Mã contest", contest.contest_id),
        ("Slug", contest.slug),
        ("Tên cuộc thi", contest.title),
        ("Trạng thái", contest.status.value if contest.status else ""),
        ("Hình thức tham gia", contest.participation_mode.value if contest.participation_mode else ""),
        ("Hình thức tổ chức", contest.delivery_mode.value if contest.delivery_mode else ""),
        ("Bắt đầu", contest.start_at.strftime("%Y-%m-%d %H:%M:%S") if contest.start_at else ""),
        ("Kết thúc", contest.end_at.strftime("%Y-%m-%d %H:%M:%S") if contest.end_at else ""),
        ("Số entries có kết quả", len(overall_rows)),
        ("Số rounds", len(rounds)),
        ("Số submissions", len(submission_rows)),
        ("Người xuất", f"{user.full_name} ({user.email})"),
        ("Thời gian xuất", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for ridx, (k, v) in enumerate(metadata_rows, start=2):
        ws4.cell(row=ridx, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=ridx, column=2, value=str(v) if v is not None else "")
    _autosize(ws4)

    # ----- Save to BytesIO + return bytes + filename -----
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"ket-qua-{contest.slug}-{today_str}.xlsx"

    return buf.read(), filename


# ============================================================
# Sprint 9b fix AD-05 (2026-05-07): system-summary xlsx export
# ============================================================

async def export_system_summary_xlsx(
    db: AsyncSession, user: AppUser, year: int
) -> tuple[bytes, str]:
    """AD-05 — Render báo cáo tổng hệ thống ra xlsx.

    Sheet 1 'Tổng quan': stat counts (users/contests/entries/submissions/certs/reviews).
    Sheet 2 'Phân loại user': breakdown theo role (STUDENT/ORGANIZER/JUDGE/HOD/ADMIN).
    Sheet 3 'Metadata': year, người xuất, thời gian xuất.

    Permission: ADMIN only — service `system_summary` enforce qua role check.
    """
    # Reuse JSON service — đã enforce admin permission + đếm stats.
    summary = await system_summary(db, user, year)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws2 = wb.create_sheet("Phân loại user")
    ws3 = wb.create_sheet("Metadata")

    # ----- Sheet 1: Tổng quan -----
    _write_headers(ws1, ["Hạng mục", "Giá trị", "Ghi chú"])
    avg_rating = summary.get("avg_review_rating")
    avg_str = f"{float(avg_rating):.2f}" if avg_rating is not None else "—"
    overview_rows = [
        ("Năm thống kê", summary["year"], "Theo created_at contest"),
        ("Tổng users (active)", summary["total_users"],
         "Đã loại DELETED status"),
        ("Tổng cuộc thi", summary["total_contests"], "Tạo trong năm"),
        ("Đã PUBLISHED+", summary["contests_published_or_after"],
         "REG_OPEN/CLOSED/ONGOING/FINISHED"),
        ("Đã FINISHED", summary["contests_finished"], "Kết thúc + đã chấm"),
        ("Tổng entries", summary["total_entries"], "Tất cả status"),
        ("Tổng bài nộp", summary["total_submissions"], "Tất cả version"),
        ("Chứng nhận đã cấp", summary["total_certificates_issued"],
         "Loại trừ revoked"),
        ("Reviews từ SV", summary["total_reviews"], "Chỉ visible"),
        ("Sao trung bình", avg_str, "Reviews visible"),
    ]
    for ridx, (k, v, note) in enumerate(overview_rows, start=2):
        ws1.cell(row=ridx, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=ridx, column=2, value=v)
        ws1.cell(row=ridx, column=3, value=note).font = Font(
            italic=True, color="6B7280")
    _autosize(ws1)

    # ----- Sheet 2: Phân loại user -----
    _write_headers(ws2, ["Vai trò", "Số lượng (active)"])
    role_rows = [
        ("STUDENT — Sinh viên", summary["students_active"]),
        ("ORGANIZER — Giảng viên BTC", summary["organizers"]),
        ("JUDGE — Giám khảo", summary["judges"]),
        ("HOD — Trưởng/CN khoa", summary["department_heads"]),
        ("ADMIN — Quản trị hệ thống", summary["admins"]),
    ]
    for ridx, (role, cnt) in enumerate(role_rows, start=2):
        ws2.cell(row=ridx, column=1, value=role)
        ws2.cell(row=ridx, column=2, value=cnt)
    _autosize(ws2)

    # ----- Sheet 3: Metadata -----
    _write_headers(ws3, ["Trường thông tin", "Giá trị"])
    metadata_rows = [
        ("Loại báo cáo", "AD-05 Tổng quan hệ thống PTIT Contest"),
        ("Năm thống kê", str(summary["year"])),
        ("Người xuất", f"{user.full_name} ({user.email})"),
        (
            "Thời gian xuất",
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        ),
    ]
    for ridx, (k, v) in enumerate(metadata_rows, start=2):
        ws3.cell(row=ridx, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=ridx, column=2, value=v)
    _autosize(ws3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"bao-cao-he-thong-{summary['year']}-{today_str}.xlsx"
    return buf.read(), filename


# ============================================================
# Sprint 23 (2026-05-09): Real-time stats endpoints
# ============================================================


async def approval_stats(
    db: AsyncSession, user, days: int = 30
) -> dict:
    """BCN-08 — Donut chart Hiệu suất duyệt N ngày qua.

    Filter theo faculty của HOD. Admin xem toàn hệ thống.
    """
    faculty_id = await _ensure_hod_or_admin(db, user)
    cutoff = datetime.now(timezone.utc) - _td(days=days)

    base = select(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(WorkflowApproval.reviewed_at >= cutoff)
    if faculty_id is not None:
        base = base.where(Contest.host_faculty_id == faculty_id)

    # Counts by status
    counts_stmt = select(
        WorkflowApproval.status, func.count()
    ).select_from(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(WorkflowApproval.reviewed_at >= cutoff)
    if faculty_id is not None:
        counts_stmt = counts_stmt.where(Contest.host_faculty_id == faculty_id)
    counts_stmt = counts_stmt.group_by(WorkflowApproval.status)

    rows = (await db.execute(counts_stmt)).all()
    counts = {st: cnt for st, cnt in rows}
    approved = counts.get(ApprovalStatus.APPROVED, 0)
    revision = counts.get(ApprovalStatus.REVISION_REQUESTED, 0)
    rejected = counts.get(ApprovalStatus.REJECTED, 0)

    # Avg processing hours (reviewed_at - submitted_at)
    avg_stmt = select(
        func.avg(
            func.extract(
                'epoch',
                WorkflowApproval.reviewed_at - WorkflowApproval.submitted_at,
            )
            / 3600.0
        )
    ).select_from(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(
        WorkflowApproval.reviewed_at >= cutoff,
        WorkflowApproval.reviewed_at.isnot(None),
    )
    if faculty_id is not None:
        avg_stmt = avg_stmt.where(Contest.host_faculty_id == faculty_id)
    avg_hours = (await db.execute(avg_stmt)).scalar()

    return {
        "days": days,
        "approved": approved,
        "revision_requested": revision,
        "rejected": rejected,
        "total": approved + revision + rejected,
        "avg_processing_hours": float(avg_hours) if avg_hours is not None else None,
    }


async def bcn_deltas(db: AsyncSession, user) -> dict:
    """BCN dashboard 4 stat cards với delta vs 24h/7d/30d trước."""
    faculty_id = await _ensure_hod_or_admin(db, user)
    now = datetime.now(timezone.utc)
    h24_ago = now - _td(hours=24)
    d7_ago = now - _td(days=7)
    d30_ago = now - _td(days=30)
    sla_24h = now + _td(hours=24)  # urgent = deadline ≤ now+24h => submitted_at ≤ now-24h (SLA 48h)
    submit_threshold = now - _td(hours=24)  # SLA 48h - 24h remain = submitted_at < now-24h

    # Queue pending (status=PENDING) hiện tại
    queue_stmt = select(func.count()).select_from(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(WorkflowApproval.status == ApprovalStatus.PENDING)
    if faculty_id is not None:
        queue_stmt = queue_stmt.where(Contest.host_faculty_id == faculty_id)
    queue_pending = (await db.execute(queue_stmt)).scalar_one()

    # Pending count 24h trước (snapshot — count items submitted before 24h ago + still pending OR reviewed > 24h ago)
    # Approximation: count approvals submitted < 24h ago + still pending
    queue_24h_stmt = select(func.count()).select_from(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(
        WorkflowApproval.submitted_at < h24_ago,
        # PENDING hoặc reviewed > h24_ago
        (WorkflowApproval.status == ApprovalStatus.PENDING) |
        (WorkflowApproval.reviewed_at > h24_ago),
    )
    if faculty_id is not None:
        queue_24h_stmt = queue_24h_stmt.where(Contest.host_faculty_id == faculty_id)
    queue_24h_ago = (await db.execute(queue_24h_stmt)).scalar_one()
    queue_delta = queue_pending - queue_24h_ago

    # Urgent: PENDING với submitted_at < submit_threshold (deadline ≤ now+24h)
    urgent_stmt = select(func.count()).select_from(WorkflowApproval).join(
        Contest, Contest.contest_id == WorkflowApproval.contest_id
    ).where(
        WorkflowApproval.status == ApprovalStatus.PENDING,
        WorkflowApproval.submitted_at < submit_threshold,
    )
    if faculty_id is not None:
        urgent_stmt = urgent_stmt.where(Contest.host_faculty_id == faculty_id)
    urgent = (await db.execute(urgent_stmt)).scalar_one()

    # Contests ongoing (REG_OPEN/REG_CLOSED/ONGOING) hiện tại
    ongoing_stmt = select(func.count()).select_from(Contest).where(
        Contest.status.in_([
            ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING
        ])
    )
    if faculty_id is not None:
        ongoing_stmt = ongoing_stmt.where(Contest.host_faculty_id == faculty_id)
    ongoing = (await db.execute(ongoing_stmt)).scalar_one()

    # Contests ongoing 7d ago: count contests có start_at <= now AND end_at > now-7d AND status was active
    # Approximation: count contests created_at <= d7_ago AND status đang active hiện tại
    ongoing_7d_stmt = select(func.count()).select_from(Contest).where(
        Contest.created_at < d7_ago,
        Contest.status.in_([
            ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING
        ])
    )
    if faculty_id is not None:
        ongoing_7d_stmt = ongoing_7d_stmt.where(Contest.host_faculty_id == faculty_id)
    ongoing_7d_ago = (await db.execute(ongoing_7d_stmt)).scalar_one()
    ongoing_delta = ongoing - ongoing_7d_ago

    # Students total: count Students có StudentDirectory.faculty_id match.
    # HOD lọc theo khoa, ADMIN xem tất.
    students_stmt = (
        select(func.count())
        .select_from(Student)
        .join(StudentDirectory,
              StudentDirectory.directory_id == Student.directory_id)
    )
    if faculty_id is not None:
        students_stmt = students_stmt.where(
            StudentDirectory.faculty_id == faculty_id
        )
    students_total = (await db.execute(students_stmt)).scalar_one()

    students_30d_stmt = (
        select(func.count())
        .select_from(Student)
        .join(StudentDirectory,
              StudentDirectory.directory_id == Student.directory_id)
        .where(Student.created_at < d30_ago)
    )
    if faculty_id is not None:
        students_30d_stmt = students_30d_stmt.where(
            StudentDirectory.faculty_id == faculty_id
        )
    students_30d_ago = (await db.execute(students_30d_stmt)).scalar_one()
    students_delta = students_total - students_30d_ago

    return {
        "queue_pending": queue_pending,
        "queue_pending_delta_24h": queue_delta,
        "urgent_count": urgent,
        "contests_ongoing": ongoing,
        "contests_ongoing_delta_7d": ongoing_delta,
        "students_total": students_total,
        "students_delta_30d": students_delta,
    }


async def btc_deltas(db: AsyncSession, user) -> dict:
    """BTC dashboard 4 stat cards với delta vs 24h/7d trước. Filter contests created_by user."""
    if "ORGANIZER" not in user.role_codes and "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ORGANIZER hoặc ADMIN")
    now = datetime.now(timezone.utc)
    h24_ago = now - _td(hours=24)
    d7_ago = now - _td(days=7)

    # Contests organizer của user
    base_where = []
    if "ADMIN" not in user.role_codes:
        base_where.append(Contest.created_by == user.user_id)

    # Ongoing
    ongoing_stmt = select(func.count()).select_from(Contest).where(
        *base_where,
        Contest.status.in_([
            ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING
        ])
    )
    ongoing = (await db.execute(ongoing_stmt)).scalar_one()

    ongoing_7d_stmt = select(func.count()).select_from(Contest).where(
        *base_where,
        Contest.created_at < d7_ago,
        Contest.status.in_([
            ContestStatus.REG_OPEN, ContestStatus.REG_CLOSED, ContestStatus.ONGOING
        ])
    )
    ongoing_7d_ago = (await db.execute(ongoing_7d_stmt)).scalar_one()

    # Sprint 28 fix Sentry (2026-05-09): Submission KHÔNG có field contest_id
    # (chỉ có entry_id + round_id). Join qua ContestEntry để lấy contest.
    # Submissions pending judge: count submissions của contests organizer
    sub_stmt = (
        select(func.count())
        .select_from(Submission)
        .join(ContestEntry, ContestEntry.entry_id == Submission.entry_id)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(
            *base_where,
            Submission.status == SubmissionStatus.SUBMITTED,
        )
    )
    submissions_pending = (await db.execute(sub_stmt)).scalar_one()

    # Submissions judged 24h: status LOCKED (đã chấm xong) updated_at >= 24h ago
    sub_24h_stmt = (
        select(func.count())
        .select_from(Submission)
        .join(ContestEntry, ContestEntry.entry_id == Submission.entry_id)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(
            *base_where,
            Submission.updated_at >= h24_ago,
            Submission.status == SubmissionStatus.LOCKED,
        )
    )
    submissions_judged_24h = (await db.execute(sub_24h_stmt)).scalar_one()

    # Registrations pending: entries PENDING của contests organizer
    reg_stmt = select(func.count()).select_from(ContestEntry).join(
        Contest, Contest.contest_id == ContestEntry.contest_id
    ).where(
        *base_where,
        ContestEntry.registration_status == RegistrationStatus.PENDING,
    )
    reg_pending = (await db.execute(reg_stmt)).scalar_one()

    reg_24h_stmt = select(func.count()).select_from(ContestEntry).join(
        Contest, Contest.contest_id == ContestEntry.contest_id
    ).where(
        *base_where,
        ContestEntry.registration_status == RegistrationStatus.PENDING,
        ContestEntry.created_at < h24_ago,
    )
    reg_pending_24h_ago = (await db.execute(reg_24h_stmt)).scalar_one()
    reg_delta = reg_pending - reg_pending_24h_ago

    # Total students unique trong các entries APPROVED+PENDING
    # (chỉ count cá nhân — bỏ team count vì team dùng team_id riêng)
    students_stmt = (
        select(func.count(func.distinct(ContestEntry.student_id)))
        .select_from(ContestEntry)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .where(
            *base_where,
            ContestEntry.registration_status.in_([
                RegistrationStatus.APPROVED, RegistrationStatus.PENDING
            ]),
            ContestEntry.student_id.isnot(None),
        )
    )
    students_total = (await db.execute(students_stmt)).scalar_one()

    return {
        "contests_ongoing": ongoing,
        "contests_ongoing_delta_7d": ongoing - ongoing_7d_ago,
        "submissions_pending_judge": submissions_pending,
        "submissions_judged_24h": submissions_judged_24h,
        "registrations_pending": reg_pending,
        "registrations_pending_delta_24h": reg_delta,
        "students_total": students_total,
    }


async def activity_feed(db: AsyncSession, user, limit: int = 10) -> dict:
    """GV-08 — Activity feed gần nhất (approve, submit, register, judge).

    Filter theo organizer của user (Admin xem tất). Lấy events:
      - WorkflowApproval submitted/reviewed
      - ContestEntry created (đăng ký)
      - Submission created (nộp bài)
    Merge sort theo timestamp desc.
    """
    if "ORGANIZER" not in user.role_codes and "ADMIN" not in user.role_codes:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần ORGANIZER hoặc ADMIN")

    is_admin = "ADMIN" in user.role_codes
    base_where_contest = [] if is_admin else [Contest.created_by == user.user_id]

    items: list[dict] = []

    # Approvals: lấy reviewed (decision) hoặc submitted (proposal) gần đây
    appr_stmt = (
        select(WorkflowApproval, Contest, AppUser)
        .join(Contest, Contest.contest_id == WorkflowApproval.contest_id)
        .join(AppUser, AppUser.user_id == WorkflowApproval.submitted_by)
        .where(*base_where_contest)
        .order_by(WorkflowApproval.updated_at.desc())
        .limit(limit)
    )
    for ap, contest, actor in (await db.execute(appr_stmt)).all():
        if ap.reviewed_at is not None:
            action = (
                "approve_q1" if ap.status == ApprovalStatus.APPROVED and ap.step.value == "BCN_QD1"
                else "approve_q2" if ap.status == ApprovalStatus.APPROVED and ap.step.value == "BCN_QD2"
                else "request_revision" if ap.status == ApprovalStatus.REVISION_REQUESTED
                else "reject" if ap.status == ApprovalStatus.REJECTED
                else "submit_proposal"
            )
            ts = ap.reviewed_at
        else:
            action = "submit_proposal"
            ts = ap.submitted_at
        items.append({
            "timestamp": ts,
            "action": action,
            "actor_name": actor.full_name,
            "contest_title": contest.title,
            "contest_id": contest.contest_id,
            "note": ap.bcn_comment if ap.reviewed_at else ap.submission_note,
        })

    # Registrations cá nhân (entries có student_id) — actor_name từ Student.user → AppUser
    reg_stmt = (
        select(ContestEntry, Contest, AppUser)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .join(Student, Student.student_id == ContestEntry.student_id)
        .join(AppUser, AppUser.user_id == Student.user_id)
        .where(*base_where_contest, ContestEntry.student_id.isnot(None))
        .order_by(ContestEntry.created_at.desc())
        .limit(limit)
    )
    for entry, contest, actor in (await db.execute(reg_stmt)).all():
        items.append({
            "timestamp": entry.created_at,
            "action": "register",
            "actor_name": actor.full_name,
            "contest_title": contest.title,
            "contest_id": contest.contest_id,
            "note": None,
        })

    # Submissions (SV nộp bài) — actor = Submission.created_by user, contest qua entry→contest
    sub_stmt = (
        select(Submission, Contest, AppUser)
        .join(ContestEntry,
              ContestEntry.entry_id == Submission.entry_id)
        .join(Contest, Contest.contest_id == ContestEntry.contest_id)
        .join(AppUser, AppUser.user_id == Submission.created_by)
        .where(*base_where_contest, Submission.created_by.isnot(None))
        .order_by(Submission.updated_at.desc())
        .limit(limit)
    )
    for submission, contest, actor in (await db.execute(sub_stmt)).all():
        # Action: SUBMITTED → "submit_work", LOCKED → "judge_locked"
        if submission.status == SubmissionStatus.LOCKED:
            action = "judge_locked"
            ts = submission.updated_at
        else:
            action = "submit_work"
            ts = submission.created_at
        items.append({
            "timestamp": ts,
            "action": action,
            "actor_name": actor.full_name,
            "contest_title": contest.title,
            "contest_id": contest.contest_id,
            "note": None,
        })

    # Sort merged + limit
    items.sort(key=lambda x: x["timestamp"], reverse=True)
    items = items[:limit]

    return {"items": items, "total": len(items)}
